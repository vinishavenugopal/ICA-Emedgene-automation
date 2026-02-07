from argparse import ArgumentParser
from bisect import bisect_left
from datetime import datetime
from openpyxl import load_workbook
from os import makedirs
from os.path import exists, join
from pandas import concat, DataFrame, ExcelWriter, isna, read_csv, Series, to_numeric
from pickle import load, dump
from sys import exit

# The options used for running this script
def create_parser():
    parser = ArgumentParser(description="Process target coverage and full resolution BED files to create panel and gene coverage files.")

    # Required arguments
    parser.add_argument("-f", "--full_res", type=str, required=True,
                        help="Path to the full resolution BED file from the 'full_res' report (<sample>.qc-coverage-region-1_full_res.bed) in ICA for a specific sample.")
    parser.add_argument("-s", "--sample_sheet", type=str, required=True,
                        help="An Illumina V2 sample sheet with the panel bed information in the \"Description\" column of the Cloud Data.")
    parser.add_argument("-n", "--sample_name", type=str, required=True,
                        help="The name of the sample so that the output goes to the correct folder and for naming of the output (<sample_name>.qc_coverage_by_level.xlsx")
    parser.add_argument("-b", "--panel_bed_folder", type=str, required=True,
                        help="The folder where all the panel bed files are located.")
    return parser

# Function to parse an Illumina V2 sample sheet
def parseSampleSheet(sampleSheetName):
    '''
    parseSampleSheet assumes that the sample sheet exists at sampleSheetName and is readable.
    The method works by reading through the sample sheet until it reaches the "Cloud_Data" line.
    Once it reaches that line, the method grabs the next line (the header) and checks for the exact
    text "Description" is the last column header under cloud data. It then reads all subsequent lines
    and grabs all the samples and their associated panel in the Description column. This means that
    the method assumes the last value in each row is associated with a panel.
    
    '''
    sampleSheetDF = DataFrame(columns=["sample","panel(s)"])
    with open(sampleSheetName) as fh:
        line = fh.readline()
        while "Cloud_Data" not in line:
            line = fh.readline()
        headerCols = fh.readline().strip('\n').strip(',').split(',')
        assert(headerCols[-1] == "Description") # The Illumina V2 sample sheet doesn't have a column named Description as the last column in the [Cloud_Data] section
        for i,line in enumerate(fh):
            if line.strip('\n').strip(',') == '': #Reached the end of the cloud data when nothing there but commas
                break
            rec = line.strip('\n').strip(',').split(',')
            if rec[0].startswith("PC") or "FILL" in rec[0]:continue
            sampleSheetDF.loc[i] = [rec[0],rec[-1]] #First column is sample id and last should be decription
    return sampleSheetDF
    
def getSampleIndex(sampleNameParts, multiSampleName):
    sampleNameParts = sampleNameParts.split('-')

    if len(sampleNameParts)>2: sampleNamePart = "-".join(sampleNameParts[-2:])
    else: sampleNamePart = sampleNameParts[-1]
    
    parts = multiSampleName.split('-')
    parts2 = parts[1].split('_')
    parts2[-1] = parts2[-1] + '-' + parts[-1]
    
    return parts2.index(sampleNamePart)

# Function to get the bed file in a DataFrame
def getPanelBed(sampleSheetDF, sampleName, panelBedFolder,multiSampleName):
    
    print("Calculating coverage metrics for",sampleName)
    cglString = str(sampleSheetDF[sampleSheetDF["sample"]==multiSampleName]["panel(s)"].values[0])
    cglIndex = 0
    if sampleName != multiSampleName:
        cglIndex = getSampleIndex(sampleName,multiSampleName)
    cgls=cglString.split('_')

    cgl = cgls[cglIndex]
    if "CGL" not in cgl and cglString.startswith("CGL"): cgl = "CGL" + cgl

    headerColumns = ["chrom","start","end","exIDs","gene"] # ,"transcript_ID","exon_number","panel"
    cglCoords = read_csv(join(panelBedFolder,cgl+".bed"),names=headerColumns,sep='\t')
    cglCoords["Panel"] = cgl

    return cglCoords,cgl

def exonCoverage(exon, coverage_data_sorted,covByCoverage):
    start_index = bisect_left(coverage_data_sorted, (exon["chrom"], exon.start))

    # Iterate forward and collect overlapping intervals
    overlapping_intervals = []
    for i in range(start_index-1, len(coverage_data_sorted)):
        chrom, start, end, coverage = coverage_data_sorted[i]
        if chrom != exon["chrom"] or start > exon.end: break # Use .chrom and .end
        if end > exon.start: overlapping_intervals.append((start, end, coverage))

    # Convert to DataFrame
    exon_coverage = DataFrame(overlapping_intervals, columns=["start", "end", "coverage"])
    
    #Get AVG coverage
    exon_coverage["dif"] = exon_coverage["end"] - exon_coverage["start"]
    totalBases = exon_coverage["dif"].sum()
    exon_coverage["prod"] = exon_coverage["coverage"] * exon_coverage["dif"]
    if exon["gene"] in covByCoverage: covByCoverage[exon["gene"]] = concat([covByCoverage[exon["gene"]],exon_coverage])
    else: covByCoverage[exon["gene"]] = exon_coverage
    if totalBases == 0:
        return Series({ 
        'AVG Coverage': 0,
        '%Bases > 0X': 0.0,
        '%Bases > 10X': 0.0,
        '%Bases > 20X': 0.0,
        '%Bases > 50X': 0.0,
        '%Bases > 100X': 0.0
    })
    
    avgCov = exon_coverage["prod"].sum()/totalBases
    percAbove_0x = exon_coverage[exon_coverage["coverage"]>0]["dif"].sum()/totalBases
    percAbove_10x = exon_coverage[exon_coverage["coverage"]>=10]["dif"].sum()/totalBases
    percAbove_20x = exon_coverage[exon_coverage["coverage"]>=20]["dif"].sum()/totalBases
    percAbove_50x = exon_coverage[exon_coverage["coverage"]>=50]["dif"].sum()/totalBases
    percAbove_100x = exon_coverage[exon_coverage["coverage"]>=100]["dif"].sum()/totalBases
    
    if exon["gene"] in covByCoverage: covByCoverage[exon["gene"]] = concat([covByCoverage[exon["gene"]],exon_coverage])
    else: covByCoverage[exon["gene"]] = exon_coverage

    return Series({ 
        'AVG Coverage': avgCov,
        '%Bases > 0X': percAbove_0x,
        '%Bases > 10X': percAbove_10x,
        '%Bases > 20X': percAbove_20x,
        '%Bases > 50X': percAbove_50x,
        '%Bases > 100X': percAbove_100x
    })

def zero():
    return Series({ 
        'AVG Coverage': 0,
        '%Bases > 0X': 0.0,
        '%Bases > 10X': 0.0,
        '%Bases > 20X': 0.0,
        '%Bases > 50X': 0.0,
        '%Bases > 100X': 0.0
    })

if __name__ == "__main__":
    parser = create_parser()
    args = parser.parse_args()

    # Access the arguments
    fullResCov = args.full_res 
    sampleName = args.sample_name 
    outName = f"{sampleName}/{sampleName}.qc_coverage_by_level.xlsx"
    sampleSheet = args.sample_sheet
    bedFolder = args.panel_bed_folder

    # Don't create reports for fill and PC samples
    if sampleName.startswith("PC") or "fill" in sampleName.lower():exit(0)

    # Check if files exist
    if not exists(sampleSheet): parser.error(f"The sample sheet '{sampleSheet}' does not exist!")
    if not exists(fullResCov): parser.error(f"The full resolution coverage file '{fullResCov}' does not exist!")

    #################### Process the target coverage report ####################
    sampleSheetDF = parseSampleSheet(sampleSheet)

    for sampleIndex,samp in enumerate(sampleName.split('_')):
        if "NGS" not in samp:
            last_two_digits = str(datetime.now().year)[-2:]
            samp = "NGS" + last_two_digits + '-' +samp
        panelBed, panelName = getPanelBed(sampleSheetDF,samp,bedFolder,sampleName)
        panelChrs = set(panelBed["chrom"].unique())
        
        #################### Process the full resolution report ####################
        # Load the full-resolution coverage BED file into a DataFrame
        print("Reading Full resolution Coverage report")
        coverage_data = read_csv(fullResCov, sep="\t", header=None, names=["chrom", "start", "end", "coverage"])

        ##################### Remove any full coverage data on chromosomes not in the panel ####################  
        print("Coverage data before filtering:",coverage_data.shape[0])
        coverage_data = coverage_data[coverage_data["chrom"].isin(panelChrs)].reset_index(drop=True)
        
        print("Coverage data after filtering:",coverage_data.shape[0])
        coverage_data_sorted = sorted(zip(coverage_data['chrom'], coverage_data['start'], coverage_data['end'], coverage_data['coverage']))
        
        #################### process the remaining lines in the full resolution report, saving all regions to a list for deeper processing ####################
        print("Calculating exon coverage.")
        covByCoverage={}
        panelBed[["AVG Coverage","%Bases > 0X","%Bases > 10X","%Bases > 20X","%Bases > 50X","%Bases > 100X"]]= panelBed.apply(lambda exon: exonCoverage(exon, coverage_data_sorted,covByCoverage), axis=1)

        #################### covByCoverage now contains genes and each gene has a list of all the exons that fall within the gene
        # the next step goes through each gene and calculates coverage and percent bases covered ####################
        panelGeneNames = panelBed["gene"].unique()
        print("Calculating gene coverage for %i genes" % (len(panelGeneNames)))
        cols = ["gene","AVG Coverage","%Bases > 0X","%Bases > 10X","%Bases > 20X","%Bases > 50X","%Bases > 100X"]
        panelGeneCov = DataFrame(columns=cols)
        allPanelCoverage = None
        for i,gene in enumerate(panelGeneNames):
            allExonCov = covByCoverage[gene]
            totalBases = allExonCov["dif"].sum()
            if totalBases == 0: continue
        
            if i == 0: allPanelCoverage = allExonCov
            else: allPanelCoverage = concat([allPanelCoverage,allExonCov])
            
            avgCov = allExonCov["prod"].sum()/totalBases
            percAbove_0x = allExonCov[allExonCov["coverage"]>0]["dif"].sum()/totalBases
            percAbove_10x = allExonCov[allExonCov["coverage"]>=10]["dif"].sum()/totalBases
            percAbove_20x = allExonCov[allExonCov["coverage"]>=20]["dif"].sum()/totalBases
            percAbove_50x = allExonCov[allExonCov["coverage"]>=50]["dif"].sum()/totalBases
            percAbove_100x = allExonCov[allExonCov["coverage"]>=100]["dif"].sum()/totalBases
            panelGeneCov.loc[panelGeneCov.shape[0],cols] = Series({ 
                'gene': gene,
                'AVG Coverage': avgCov,
                '%Bases > 0X': percAbove_0x,
                '%Bases > 10X': percAbove_10x,
                '%Bases > 20X': percAbove_20x,
                '%Bases > 50X': percAbove_50x,
                '%Bases > 100X': percAbove_100x
            })
        
        #################### The previous step saved all the genes in the panel to the same list of exons. Now this step will process all the exons together ####################
        print("Calculating panel coverage")
        cols = ["Panel","AVG Coverage","%Bases > 0X","%Bases > 10X","%Bases > 20X","%Bases > 50X","%Bases > 100X"]
        panelCov = DataFrame(columns=cols)
        
        totalBases = allPanelCoverage["dif"].sum()
        avgCov = allPanelCoverage["prod"].sum()/totalBases
        percAbove_0x = allPanelCoverage[allPanelCoverage["coverage"]>0]["dif"].sum()/totalBases
        percAbove_10x = allPanelCoverage[allPanelCoverage["coverage"]>=10]["dif"].sum()/totalBases
        percAbove_20x = allPanelCoverage[allPanelCoverage["coverage"]>=20]["dif"].sum()/totalBases
        percAbove_50x = allPanelCoverage[allPanelCoverage["coverage"]>=50]["dif"].sum()/totalBases
        percAbove_100x = allPanelCoverage[allPanelCoverage["coverage"]>=100]["dif"].sum()/totalBases
        
        panelCov.loc[0,cols] = Series(
            {'Panel': panelName,
            'AVG Coverage': avgCov,
            '%Bases > 0X': percAbove_0x,
            '%Bases > 10X': percAbove_10x,
            '%Bases > 20X': percAbove_20x,
            '%Bases > 50X': percAbove_50x,
            '%Bases > 100X': percAbove_100x}
        )
        
        #################### Step to ensure the results are saved to an existing directory containing the sample ####################
        makedirs(sampleName, exist_ok=True)

        print(outName)
        percent_cols = ['%Bases > 0X', '%Bases > 10X', '%Bases > 20X', '%Bases > 50X', '%Bases > 100X']
        if exists(outName):
            # Load the existing workbook
            book = load_workbook(outName)
        
            with ExcelWriter(outName, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Automatically maps existing sheets (no need to set `writer.book`)
                panelBed["Panel"] = panelName
                panelGeneCov["Panel"] = panelName
        
                # Find the last row in each sheet
                startrow_exon = book['Exon Coverage'].max_row
                startrow_gene = book['Gene Coverage'].max_row
                startrow_panel = book['Panel Coverage'].max_row
        
                # Append data to existing sheets
                panelBed.to_excel(writer, sheet_name='Exon Coverage', index=None, startrow=startrow_exon, header=False)
                panelGeneCov.to_excel(writer, sheet_name='Gene Coverage', index=None, startrow=startrow_gene, header=False)
                panelCov.to_excel(writer, sheet_name='Panel Coverage', index=None, startrow=startrow_panel, header=False)
        
            # Reopen workbook to apply formatting
            book = load_workbook(outName)
        
            for sheet_name, df in zip(['Panel Coverage', 'Gene Coverage', 'Exon Coverage'], [panelCov, panelGeneCov, panelBed]):
                worksheet = book[sheet_name]
                col_indices = [df.columns.get_loc(col) + 1 for col in percent_cols]  # +1 to match Excel's 1-based columns
        
                for col_num in col_indices:
                    for row in worksheet.iter_rows(min_row=2, min_col=col_num, max_col=col_num):  # Skip header
                        for cell in row:
                            cell.number_format = '0%'  # Apply percentage format
        
            book.save(outName)  # Save changes
        
        else:
            # Create a new Excel file if it doesn't exist
            with ExcelWriter(outName, engine='openpyxl', mode='w') as writer:
                panelBed["Panel"] = panelName
                panelGeneCov["Panel"] = panelName
                panelBed.to_excel(writer, sheet_name='Exon Coverage', index=None)
                panelGeneCov.to_excel(writer, sheet_name='Gene Coverage', index=None)
                panelCov.to_excel(writer, sheet_name='Panel Coverage', index=None)
        
            # Apply formatting after saving
            book = load_workbook(outName)
        
            for sheet_name, df in zip(['Panel Coverage', 'Gene Coverage', 'Exon Coverage'], [panelCov, panelGeneCov, panelBed]):
                worksheet = book[sheet_name]
                col_indices = [df.columns.get_loc(col) + 1 for col in percent_cols]
        
                for col_num in col_indices:
                    for row in worksheet.iter_rows(min_row=2, min_col=col_num, max_col=col_num):  # Skip header
                        for cell in row:
                            cell.number_format = '0%'
        
            book.save(outName)  # Save changes




